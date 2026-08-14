#!/usr/bin/env python3
"""Собирает catalog.json из таблицы каталога.

Источник — catalog/catalog.xlsx (если есть и установлен openpyxl), иначе catalog/catalog.csv.
Результат кладётся в ios/KBZHU/Resources/catalog.json: он же уезжает на CDN и он же
лежит в бандле приложения как стартовый снимок.

Запуск:  python3 tools/build_catalog.py [--check]
  --check  ничего не пишет, только проверяет таблицу и то, что JSON не устарел.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from normalize_brands import normalize_key

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "catalog" / "catalog.csv"
XLSX_PATH = ROOT / "catalog" / "catalog.xlsx"
OUT_PATH = ROOT / "ios" / "KBZHU" / "Resources" / "catalog.json"

# Единицы, которые понимает приложение. Пустая — продукт считается в граммах.
UNITS = {"порция", "шт", "стакан", "пачка", "баночка", "ломтик"}

# Типы позиций. Пусто — обычная упаковка; «ресторан» — блюдо сети питания.
KINDS = {"ресторан"}

# Заголовки колонок: канон -> допустимые написания.
COLUMNS = {
    "name": {"название", "наименование", "продукт", "name"},
    "manufacturer": {"производитель", "бренд", "manufacturer", "brand"},
    "kcal": {"ккал_100г", "ккал", "калории", "kcal"},
    "protein": {"белки_100г", "белки", "белок", "protein"},
    "fat": {"жиры_100г", "жиры", "fat"},
    "carbs": {"углеводы_100г", "углеводы", "carbs"},
    "unit": {"единица", "unit"},
    "unit_grams": {"вес_единицы_г", "вес_единицы", "вес_порции_г", "unit_grams"},
    "default_grams": {"порция_по_умолчанию_г", "порция_по_умолчанию", "default_grams"},
    "barcode": {"штрихкод", "штрих-код", "barcode"},
    "id": {"id", "ид"},
    "kind": {"тип", "kind", "type"},
}

REQUIRED = ("name", "kcal", "protein", "fat", "carbs")

TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh",
    "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "c",
    "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e",
    "ю": "yu", "я": "ya",
}


def slugify(text: str) -> str:
    out = []
    for ch in text.lower():
        if ch in TRANSLIT:
            out.append(TRANSLIT[ch])
        elif ch.isalnum():
            out.append(ch)
        else:
            out.append("-")
    slug = re.sub(r"-+", "-", "".join(out)).strip("-")
    return slug or "product"


def normalize_header(raw: str) -> str | None:
    key = (raw or "").strip().lower().replace(" ", "_").lstrip("﻿")
    for canon, aliases in COLUMNS.items():
        if key == canon or key in aliases:
            return canon
    return None


def read_rows() -> tuple[list[dict], str]:
    if XLSX_PATH.exists():
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Найден catalog.xlsx, но нет openpyxl. Поставьте: pip install openpyxl")
        sheet = load_workbook(XLSX_PATH, data_only=True).active
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            sys.exit("catalog.xlsx пустой")
        headers = [normalize_header(str(h or "")) for h in raw_rows[0]]
        rows = []
        for values in raw_rows[1:]:
            row = {}
            for header, value in zip(headers, values):
                if header:
                    row[header] = "" if value is None else str(value).strip()
            rows.append(row)
        return rows, XLSX_PATH.name

    if not CSV_PATH.exists():
        sys.exit(f"Не найден ни {XLSX_PATH.name}, ни {CSV_PATH.name}")

    with CSV_PATH.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            headers = [normalize_header(h) for h in next(reader)]
        except StopIteration:
            sys.exit("catalog.csv пустой")
        rows = []
        for values in reader:
            row = {}
            for header, value in zip(headers, values):
                if header:
                    row[header] = (value or "").strip()
            rows.append(row)
    return rows, CSV_PATH.name


def number(raw: str, field: str, line: int, errors: list[str], required: bool) -> float | None:
    text = (raw or "").replace(",", ".").replace(" ", "").replace(" ", "")
    if not text:
        if required:
            errors.append(f"строка {line}: не заполнено «{field}»")
        return None
    try:
        value = float(text)
    except ValueError:
        errors.append(f"строка {line}: «{field}» — не число ({raw!r})")
        return None
    if value < 0:
        errors.append(f"строка {line}: «{field}» отрицательное")
        return None
    return value


def build() -> dict:
    rows, source = read_rows()
    errors: list[str] = []
    products: list[dict] = []
    seen_ids: dict[str, int] = {}
    seen_barcodes: dict[str, int] = {}
    brand_spellings: dict[str, set[str]] = {}

    for index, row in enumerate(rows):
        line = index + 2  # +1 заголовок, +1 нумерация с единицы
        if not any((value or "").strip() for value in row.values()):
            continue  # пустая строка в конце таблицы — не ошибка

        name = (row.get("name") or "").strip()
        if not name:
            errors.append(f"строка {line}: пустое название")
            continue

        unit = (row.get("unit") or "").strip().lower()
        if unit and unit not in UNITS:
            errors.append(f"строка {line}: единица «{unit}» неизвестна, допустимы: {', '.join(sorted(UNITS))}")
            continue

        kind = (row.get("kind") or "").strip().lower()
        if kind and kind not in KINDS:
            errors.append(f"строка {line}: тип «{kind}» неизвестен, допустимы: {', '.join(sorted(KINDS))}")
            continue

        macros = {
            field: number(row.get(field, ""), field, line, errors, required=True)
            for field in ("kcal", "protein", "fat", "carbs")
        }
        if any(value is None for value in macros.values()):
            continue

        unit_grams = number(row.get("unit_grams", ""), "вес_единицы_г", line, errors, required=bool(unit))
        if unit and (unit_grams is None or unit_grams <= 0):
            if unit_grams is not None:
                errors.append(f"строка {line}: у единицы «{unit}» нужен вес больше нуля")
            continue

        default_grams = number(row.get("default_grams", ""), "порция_по_умолчанию_г", line, errors, required=False)
        if default_grams is None or default_grams <= 0:
            default_grams = unit_grams if unit_grams else 100.0

        manufacturer = (row.get("manufacturer") or "").strip()
        if manufacturer:
            brand_spellings.setdefault(normalize_key(manufacturer), set()).add(manufacturer)
        barcode = re.sub(r"\s+", "", row.get("barcode") or "")
        if barcode and not barcode.isdigit():
            errors.append(f"строка {line}: штрихкод должен состоять только из цифр")
            continue

        product_id = (row.get("id") or "").strip() or slugify(f"{name}-{manufacturer}")

        if product_id in seen_ids:
            errors.append(
                f"строка {line}: id «{product_id}» уже занят строкой {seen_ids[product_id]}. "
                "Задайте разные id в колонке id."
            )
            continue
        seen_ids[product_id] = line

        if barcode:
            if barcode in seen_barcodes:
                errors.append(f"строка {line}: штрихкод {barcode} уже есть в строке {seen_barcodes[barcode]}")
                continue
            seen_barcodes[barcode] = line

        product = {
            "id": product_id,
            "name": name,
            "manufacturer": manufacturer,
            "kcal": macros["kcal"],
            "protein": macros["protein"],
            "fat": macros["fat"],
            "carbs": macros["carbs"],
            "defaultGrams": default_grams,
        }
        if unit:
            product["unit"] = unit
            product["unitGrams"] = unit_grams
        if barcode:
            product["barcode"] = barcode
        if kind:
            product["kind"] = kind
        products.append(product)

    # Один производитель — одно написание. Иначе варианты наползут снова
    # при следующем импорте из Open Food Facts.
    for spellings in brand_spellings.values():
        if len(spellings) > 1:
            listed = "», «".join(sorted(spellings))
            errors.append(
                f"производитель записан по-разному: «{listed}». "
                "Сведите к одному написанию: tools/normalize_brands.py"
            )

    if errors:
        print(f"Каталог не собран, ошибок: {len(errors)}", file=sys.stderr)
        for error in errors:
            print("  • " + error, file=sys.stderr)
        sys.exit(1)

    payload = json.dumps(products, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    revision = hashlib.sha1(payload.encode("utf-8")).hexdigest()[:12]
    print(f"{source}: {len(products)} продуктов, ревизия {revision}")

    return {
        "revision": revision,
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "products": products,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true", help="только проверить, ничего не писать")
    args = parser.parse_args()

    catalog = build()
    text = json.dumps(catalog, ensure_ascii=False, indent=2) + "\n"

    if args.check:
        if not OUT_PATH.exists():
            sys.exit(f"{OUT_PATH.relative_to(ROOT)} отсутствует — запустите build_catalog.py")
        current = json.loads(OUT_PATH.read_text(encoding="utf-8"))
        if current.get("revision") != catalog["revision"]:
            sys.exit(
                f"{OUT_PATH.relative_to(ROOT)} устарел: в таблице ревизия {catalog['revision']}, "
                f"в файле {current.get('revision')}. Запустите build_catalog.py и закоммитьте результат."
            )
        print("Каталог актуален")
        return

    # generatedAt меняется каждый запуск — не переписываем файл, если данные те же.
    if OUT_PATH.exists():
        current = json.loads(OUT_PATH.read_text(encoding="utf-8"))
        if current.get("revision") == catalog["revision"]:
            print("Без изменений")
            return

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(text, encoding="utf-8")
    print(f"Записан {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
